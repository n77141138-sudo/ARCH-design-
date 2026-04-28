"""
Agent Ana - 報價精算系統 [Quote_Engine Skill]
功能：掃描廠商報價單資料夾 → 彙整成本 → 依毛利比例產出對業主報價 Excel + Markdown 表格

利潤公式：客戶報價 = 廠商成本 × (1 + 利潤率)
預設利潤率：20%（markup_rate: 1.2），可在 config.json 調整
輸出格式：Excel（審核用彙整表 + 對業主報價單）+ Markdown 表格（可直接貼入 Excel）
"""

import os
import glob
import json
import re
from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter


# ─── 路徑設定 ─────────────────────────────────────────────
BASE_DIR = Path(__file__).parent
INPUT_DIR  = BASE_DIR / "01_vendor_quotes_input"
OUTPUT_DIR = BASE_DIR / "02_client_quotes_output"
REVIEW_DIR = BASE_DIR / "03_review_summaries"

for d in [INPUT_DIR, OUTPUT_DIR, REVIEW_DIR]:
    d.mkdir(parents=True, exist_ok=True)


# ─── 顏色/樣式常數 ────────────────────────────────────────
COLOR_HEADER_DARK   = "1F3864"   # 深藍（主標題）
COLOR_HEADER_MID    = "2E75B6"   # 中藍（欄位標題）
COLOR_SECTION       = "BDD7EE"   # 淺藍（分類列）
COLOR_SUBTOTAL      = "D6E4F0"   # 極淺藍（小計列）
COLOR_TOTAL         = "1F3864"   # 深藍（總計列）
COLOR_WHITE         = "FFFFFF"
COLOR_ALT_ROW       = "F2F9FF"   # 交替行底色

FONT_NAME = "微軟正黑體"


def make_border(style: str = "thin") -> Border:
    s = Side(border_style=style, color="B8CCE4")
    return Border(left=s, right=s, top=s, bottom=s)


def apply_header_style(cell, bg_color: str, font_size: int = 11,
                        bold: bool = True, font_color: str = "FFFFFF",
                        wrap: bool = False):
    cell.font = Font(name=FONT_NAME, size=font_size, bold=bold, color=font_color)
    cell.fill = PatternFill("solid", fgColor=bg_color)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=wrap)
    cell.border = make_border()


def apply_data_style(cell, bold: bool = False, align: str = "center",
                      bg_color: str = COLOR_WHITE, number_format: str = None):
    cell.font = Font(name=FONT_NAME, size=10, bold=bold, color="1A1A2E")
    cell.fill = PatternFill("solid", fgColor=bg_color)
    cell.alignment = Alignment(horizontal=align, vertical="center")
    cell.border = make_border()
    if number_format:
        cell.number_format = number_format


# ─── 資料讀取：手動輸入的 JSON 廠商資料 ─────────────────────
def load_vendor_json(json_path: Path) -> list[dict]:
    """
    讀取廠商報價 JSON 格式：
    [
      {
        "vendor": "廠商名稱",
        "category": "工程類別",
        "items": [
          {"name": "項目名稱", "spec": "規格", "qty": 1, "unit": "式", "unit_price": 10000}
        ]
      }
    ]
    """
    with open(json_path, encoding="utf-8") as f:
        return json.load(f)


def load_vendor_excel(xlsx_path: Path) -> list[dict]:
    """
    讀取廠商報價 Excel（需符合欄位：項目, 規格, 數量, 單位, 單價）
    """
    df = pd.read_excel(xlsx_path)
    df.columns = [str(c).strip() for c in df.columns]

    col_map = {
        "項目": "name", "品名": "name", "品名規格": "name",
        "規格": "spec", "備註": "spec",
        "數量": "qty", "數 量": "qty",
        "單位": "unit", "單 位": "unit",
        "單價": "unit_price", "單 價": "unit_price",
    }
    df.rename(columns={k: v for k, v in col_map.items() if k in df.columns}, inplace=True)

    required = ["name", "qty", "unit_price"]
    for col in required:
        if col not in df.columns:
            df[col] = ""

    df["spec"]  = df.get("spec", "")
    df["unit"]  = df.get("unit", "式")
    df = df.dropna(subset=["name"]).fillna("")

    vendor_name = xlsx_path.stem
    category    = xlsx_path.parent.name if xlsx_path.parent.name != "01_vendor_quotes_input" else "其他"

    items = []
    for _, row in df.iterrows():
        try:
            qty   = float(row["qty"])   if str(row["qty"]).strip()        != "" else 1
            price = float(str(row["unit_price"]).replace(",", "")) if str(row["unit_price"]).strip() != "" else 0
        except ValueError:
            continue
        if price == 0:
            continue
        items.append({
            "name":       str(row["name"]).strip(),
            "spec":       str(row.get("spec", "")).strip(),
            "qty":        qty,
            "unit":       str(row.get("unit", "式")).strip() or "式",
            "unit_price": price,
        })

    return [{"vendor": vendor_name, "category": category, "items": items}]


def scan_input_folder() -> list[dict]:
    """掃描輸入資料夾，讀取所有支援的廠商報價檔案"""
    all_vendors = []

    # 掃描 JSON
    for json_file in INPUT_DIR.rglob("*.json"):
        try:
            all_vendors.extend(load_vendor_json(json_file))
            print(f"  [OK] 已讀取 JSON: {json_file.name}")
        except Exception as e:
            print(f"  [ERR] 讀取失敗 {json_file.name}: {e}")

    # 掃描 Excel（排除暫存檔 ~$）
    for xlsx_file in INPUT_DIR.rglob("*.xlsx"):
        if xlsx_file.name.startswith("~$"):
            continue
        try:
            all_vendors.extend(load_vendor_excel(xlsx_file))
            print(f"  [OK] 已讀取 Excel: {xlsx_file.name}")
        except Exception as e:
            print(f"  [ERR] 讀取失敗 {xlsx_file.name}: {e}")

    return all_vendors


# ─── 毛利計算 ─────────────────────────────────────────────
def calculate_client_price(cost_price: float, markup_rate: float) -> float:
    """markup_rate: 例如 1.3 代表加價 30%"""
    return round(cost_price * markup_rate, 0)


# ─── 產出審核用彙整 Excel ─────────────────────────────────
def export_review_excel(
    vendors: list[dict],
    markup_rate: float,
    project_name: str,
    output_path: Path
):
    wb = Workbook()
    ws = wb.active
    ws.title = "彙整審核表"

    # ── 標題列
    ws.merge_cells("A1:I1")
    title_cell = ws["A1"]
    title_cell.value = f"【JYY DESIGN】廠商報價彙整審核表｜{project_name}"
    title_cell.font  = Font(name=FONT_NAME, size=14, bold=True, color=COLOR_WHITE)
    title_cell.fill  = PatternFill("solid", fgColor=COLOR_HEADER_DARK)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 32

    ws.merge_cells("A2:I2")
    info_cell = ws["A2"]
    info_cell.value = (
        f"產出時間：{datetime.now().strftime('%Y/%m/%d %H:%M')}　｜　"
        f"加價倍率：{markup_rate}x（利潤 {(markup_rate-1)*100:.0f}%）　｜　"
        f"請確認每項金額後再輸出對業主報價單"
    )
    info_cell.font      = Font(name=FONT_NAME, size=10, color="2E75B6")
    info_cell.fill      = PatternFill("solid", fgColor="EBF5FB")
    info_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 20

    # ── 欄位標題
    headers = ["#", "工程類別", "廠商", "項目名稱", "規格/備註", "數量", "單位",
               "廠商單價(成本)", "對客單價", "廠商小計", "對客小計", "加價倍率"]
    col_widths = [5, 14, 14, 30, 25, 7, 6, 14, 14, 14, 14, 10]

    for col_idx, (header, width) in enumerate(zip(headers, col_widths), start=1):
        cell = ws.cell(row=3, column=col_idx, value=header)
        apply_header_style(cell, COLOR_HEADER_MID, font_size=10)
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.row_dimensions[3].height = 22

    # ── 填入資料
    row_num     = 4
    item_counter = 1
    grand_cost   = 0.0
    grand_client = 0.0

    # 依 category 分組
    from itertools import groupby
    all_items_flat = []
    for vendor in vendors:
        for item in vendor["items"]:
            all_items_flat.append({
                "category":   vendor.get("category", "其他"),
                "vendor":     vendor.get("vendor",   "未知"),
                "name":       item["name"],
                "spec":       item.get("spec", ""),
                "qty":        item["qty"],
                "unit":       item.get("unit", "式"),
                "unit_price": item["unit_price"],
                "markup":     item.get("markup", markup_rate),
            })

    # 分類排序
    all_items_flat.sort(key=lambda x: x["category"])

    current_category = None
    cat_cost_sum  = 0.0
    cat_client_sum = 0.0
    cat_start_row  = None

    def write_category_subtotal():
        nonlocal row_num
        ws.merge_cells(f"A{row_num}:G{row_num}")
        sub_label = ws.cell(row=row_num, column=1,
                            value=f"► {current_category} 小計")
        sub_label.font  = Font(name=FONT_NAME, size=10, bold=True, color="1F3864")
        sub_label.fill  = PatternFill("solid", fgColor=COLOR_SUBTOTAL)
        sub_label.alignment = Alignment(horizontal="right", vertical="center")

        for c_off, val in enumerate([cat_cost_sum, cat_client_sum], start=0):
            c = ws.cell(row=row_num, column=10 + c_off, value=val)
            apply_data_style(c, bold=True, bg_color=COLOR_SUBTOTAL,
                             number_format="#,##0")
        ws.cell(row=row_num, column=8).fill = PatternFill("solid", fgColor=COLOR_SUBTOTAL)
        ws.cell(row=row_num, column=9).fill = PatternFill("solid", fgColor=COLOR_SUBTOTAL)
        ws.cell(row=row_num, column=12).fill = PatternFill("solid", fgColor=COLOR_SUBTOTAL)
        ws.row_dimensions[row_num].height = 18
        row_num += 1

    for idx, item in enumerate(all_items_flat):
        cat = item["category"]

        # 分類標題列
        if cat != current_category:
            if current_category is not None:
                write_category_subtotal()
                cat_cost_sum   = 0.0
                cat_client_sum = 0.0

            current_category = cat
            ws.merge_cells(f"A{row_num}:L{row_num}")
            cat_cell = ws.cell(row=row_num, column=1, value=f"▌ {cat}")
            apply_header_style(cat_cell, COLOR_SECTION, font_size=10,
                               font_color="1F3864", bold=True)
            ws.row_dimensions[row_num].height = 20
            row_num += 1

        # 計算價格
        client_up = calculate_client_price(item["unit_price"], item["markup"])
        cost_total   = item["unit_price"] * item["qty"]
        client_total = client_up * item["qty"]

        cat_cost_sum   += cost_total
        cat_client_sum += client_total
        grand_cost     += cost_total
        grand_client   += client_total

        # 交替行底色
        bg = COLOR_ALT_ROW if item_counter % 2 == 0 else COLOR_WHITE

        data_row = [
            item_counter,
            item["category"],
            item["vendor"],
            item["name"],
            item["spec"],
            item["qty"],
            item["unit"],
            item["unit_price"],
            client_up,
            cost_total,
            client_total,
            f"x{item['markup']}",
        ]
        for col_idx, val in enumerate(data_row, start=1):
            cell = ws.cell(row=row_num, column=col_idx, value=val)
            align = "left" if col_idx in [3, 4, 5] else "center"
            fmt   = "#,##0" if col_idx in [8, 9, 10, 11] else None
            apply_data_style(cell, align=align, bg_color=bg, number_format=fmt)

        ws.row_dimensions[row_num].height = 18
        row_num     += 1
        item_counter += 1

    # 最後一個分類小計
    if current_category is not None:
        write_category_subtotal()

    # ── 總計列
    ws.merge_cells(f"A{row_num}:G{row_num}")
    total_label = ws.cell(row=row_num, column=1, value="▶ 總計合計")
    total_label.font  = Font(name=FONT_NAME, size=11, bold=True, color=COLOR_WHITE)
    total_label.fill  = PatternFill("solid", fgColor=COLOR_TOTAL)
    total_label.alignment = Alignment(horizontal="right", vertical="center")

    for c_off, val in enumerate([grand_cost, grand_client], start=0):
        c = ws.cell(row=row_num, column=10 + c_off, value=val)
        c.font      = Font(name=FONT_NAME, size=11, bold=True, color=COLOR_WHITE)
        c.fill      = PatternFill("solid", fgColor=COLOR_TOTAL)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.number_format = "#,##0"
        c.border    = make_border()
    for c_idx in [8, 9, 12]:
        ws.cell(row=row_num, column=c_idx).fill = PatternFill("solid", fgColor=COLOR_TOTAL)

    ws.row_dimensions[row_num].height = 26
    row_num += 1

    # ── 凍結標題
    ws.freeze_panes = "A4"

    wb.save(output_path)
    print(f"  [SAVED] 審核表：{output_path.name}")
    return grand_cost, grand_client


# ─── 產出對業主正式報價 Excel ──────────────────────────────
def export_client_quote_excel(
    vendors: list[dict],
    markup_rate: float,
    project_name: str,
    project_address: str,
    quote_date: str,
    validity_days: int,
    output_path: Path,
    project_category: str = ""
):
    wb = Workbook()
    ws = wb.active
    ws.title = "對業主報價單"

    # 欄寬
    col_widths = [6, 35, 28, 8, 6, 14, 14]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # ── 公司標題區（rows 1-5）
    ws.merge_cells("A1:G1")
    h1 = ws["A1"]
    h1.value     = "工程估價單"
    h1.font      = Font(name=FONT_NAME, size=18, bold=True, color=COLOR_WHITE)
    h1.fill      = PatternFill("solid", fgColor=COLOR_HEADER_DARK)
    h1.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 40

    info_pairs = [
        ("案件名稱：", project_name, "案件地址：", project_address),
        ("報價日期：", quote_date, "有效期限：", f"{validity_days} 天"),
    ]
    for r_offset, (l1, v1, l2, v2) in enumerate(info_pairs, start=2):
        ws.merge_cells(f"A{r_offset}:B{r_offset}")
        ws.merge_cells(f"C{r_offset}:D{r_offset}")
        ws.merge_cells(f"E{r_offset}:G{r_offset}")

        label1 = ws.cell(row=r_offset, column=1, value=l1 + v1)
        label1.font      = Font(name=FONT_NAME, size=10, bold=False, color="1A1A2E")
        label1.fill      = PatternFill("solid", fgColor="EBF5FB")
        label1.alignment = Alignment(horizontal="left", vertical="center")
        label1.border    = make_border()

        label2 = ws.cell(row=r_offset, column=3, value=l2 + v2)
        label2.font      = Font(name=FONT_NAME, size=10, color="1A1A2E")
        label2.fill      = PatternFill("solid", fgColor="EBF5FB")
        label2.alignment = Alignment(horizontal="left", vertical="center")
        label2.border    = make_border()

        ws.row_dimensions[r_offset].height = 18

    # ── 備用空白列
    ws.row_dimensions[4].height = 6

    # ── 欄位標題（row 5）
    headers = ["#", "項目名稱", "規格/備註", "數量", "單位", "單價", "總計"]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=5, column=col_idx, value=header)
        apply_header_style(cell, COLOR_HEADER_MID, font_size=10)
    ws.row_dimensions[5].height = 22

    # ── 資料列
    row_num      = 6
    item_counter = 1
    grand_total  = 0.0

    all_items_flat = []
    for vendor in vendors:
        for item in vendor["items"]:
            all_items_flat.append({
                "category":   vendor.get("category", "其他"),
                "name":       item["name"],
                "spec":       item.get("spec", ""),
                "qty":        item["qty"],
                "unit":       item.get("unit", "式"),
                "unit_price": item["unit_price"],
                "markup":     item.get("markup", markup_rate),
            })

    all_items_flat.sort(key=lambda x: x["category"])
    current_category = None

    for item in all_items_flat:
        cat = item["category"]

        if cat != current_category:
            current_category = cat
            ws.merge_cells(f"A{row_num}:G{row_num}")
            cat_cell = ws.cell(row=row_num, column=1, value=f"▌ {cat}")
            apply_header_style(cat_cell, COLOR_SECTION, font_size=10,
                               font_color="1F3864", bold=True)
            ws.row_dimensions[row_num].height = 20
            row_num += 1

        client_up    = calculate_client_price(item["unit_price"], item["markup"])
        client_total = client_up * item["qty"]
        grand_total += client_total

        bg = COLOR_ALT_ROW if item_counter % 2 == 0 else COLOR_WHITE
        row_data = [item_counter, item["name"], item["spec"],
                    item["qty"], item["unit"], client_up, client_total]

        for col_idx, val in enumerate(row_data, start=1):
            cell = ws.cell(row=row_num, column=col_idx, value=val)
            align = "left" if col_idx in [2, 3] else "center"
            fmt   = "#,##0" if col_idx in [6, 7] else None
            apply_data_style(cell, align=align, bg_color=bg, number_format=fmt)

        ws.row_dimensions[row_num].height = 18
        row_num     += 1
        item_counter += 1

    # ── 總計
    row_num += 1
    
    # 根據分類決定是否收取工程管理費
    if project_category in ["室內設計", "預售客變"]:
        management_fee = 0
    else:
        management_fee = grand_total * 0.1
        
    final_total = grand_total + management_fee
    
    # 1. 合計金額
    ws.merge_cells(f"A{row_num}:F{row_num}")
    ws.cell(row=row_num, column=1, value="合計金額").alignment = Alignment(horizontal="right")
    ws.cell(row=row_num, column=7, value=grand_total).number_format = "#,##0"
    row_num += 1
    
    # 2. 工程管理費 (如果有)
    if management_fee > 0:
        ws.merge_cells(f"A{row_num}:F{row_num}")
        ws.cell(row=row_num, column=1, value="工程管理費 (10%)").alignment = Alignment(horizontal="right")
        ws.cell(row=row_num, column=7, value=management_fee).number_format = "#,##0"
        row_num += 1
    
    # 3. 總計金額
    ws.merge_cells(f"A{row_num}:F{row_num}")
    total_label = ws.cell(row=row_num, column=1, value="總計金額")
    total_label.font      = Font(name=FONT_NAME, size=11, bold=True, color=COLOR_WHITE)
    total_label.fill      = PatternFill("solid", fgColor=COLOR_TOTAL)
    total_label.alignment = Alignment(horizontal="right", vertical="center")
    total_label.border    = make_border()

    total_val = ws.cell(row=row_num, column=7, value=final_total)
    total_val.font      = Font(name=FONT_NAME, size=12, bold=True, color=COLOR_WHITE)
    total_val.fill      = PatternFill("solid", fgColor=COLOR_TOTAL)
    total_val.alignment = Alignment(horizontal="center", vertical="center")
    total_val.number_format = "#,##0"
    total_val.border    = make_border()
    ws.row_dimensions[row_num].height = 26

    # ── 備註列
    row_num += 2
    ws.merge_cells(f"A{row_num}:G{row_num}")
    
    if management_fee > 0:
        note_text = "※ 以上報價均為新台幣，內含10%工程管理費用。實際施工以現場丈量確認為主。"
    else:
        note_text = "※ 以上報價均為新台幣。實際施工以現場丈量確認為主。"
        
    note = ws.cell(row=row_num, column=1, value=note_text)
    note.font      = Font(name=FONT_NAME, size=9, color="888888")
    note.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[row_num].height = 16

    ws.freeze_panes = "A6"
    wb.save(output_path)
    print(f"  [SAVED] 報價單：{output_path.name}")
    return grand_total


# ─── 產出 Markdown 表格（Quote_Engine Skill 標準輸出）───────
def export_markdown_quote(
    vendors: list[dict],
    markup_rate: float,
    project_name: str,
    project_address: str,
    quote_date: str,
    validity_days: int,
    output_path: Path,
    project_category: str = ""
) -> str:
    """
    輸出標準 Markdown 表格報價單。
    公式：客戶報價 = 廠商成本 × (1 + 利潤率)
    利潤率 = markup_rate - 1，例如 markup_rate=1.2 → 利潤率 20%
    回傳 Markdown 字串，同時寫入 .md 檔案。
    """
    profit_pct = (markup_rate - 1) * 100

    lines: list[str] = []
    lines.append(f"# 工程估價單｜{project_name}")
    lines.append(f"")
    lines.append(f"| 資訊 | 內容 |")
    lines.append(f"| :--- | :--- |")
    lines.append(f"| 案件名稱 | {project_name} |")
    lines.append(f"| 案件地址 | {project_address} |")
    lines.append(f"| 報價日期 | {quote_date} |")
    lines.append(f"| 有效期限 | {validity_days} 天 |")
    lines.append(f"")
    lines.append(f"---")
    lines.append(f"")

    # 依類別整理
    all_items_flat: list[dict] = []
    for vendor in vendors:
        for item in vendor["items"]:
            all_items_flat.append({
                "category":   vendor.get("category", "其他"),
                "vendor":     vendor.get("vendor",   "未知"),
                "name":       item["name"],
                "spec":       item.get("spec", ""),
                "qty":        item["qty"],
                "unit":       item.get("unit", "式"),
                "unit_price": item["unit_price"],
                "markup":     item.get("markup", markup_rate),
            })
    all_items_flat.sort(key=lambda x: x["category"])

    current_category = None
    item_counter     = 1
    grand_total      = 0.0
    cat_total        = 0.0

    def open_table():
        lines.append("| # | 項目名稱 | 規格/備註 | 數量 | 單位 | 單價 | 總計 |")
        lines.append("| :---: | :--- | :--- | :---: | :---: | ---: | ---: |")

    def close_table(category: str, subtotal: float):
        lines.append(f"| | **▶ {category} 小計** | | | | | **NT$ {subtotal:,.0f}** |")
        lines.append(f"")

    for item in all_items_flat:
        cat = item["category"]

        if cat != current_category:
            # 關閉上一個類別
            if current_category is not None:
                close_table(current_category, cat_total)
            current_category = cat
            cat_total        = 0.0

            lines.append(f"## {cat}")
            lines.append(f"")
            open_table()

        # 計算價格：客戶報價 = 廠商成本 × (1 + 利潤率)
        client_up    = round(item["unit_price"] * item["markup"], 0)
        client_total = client_up * item["qty"]
        cat_total   += client_total
        grand_total += client_total

        spec_text = item["spec"] if item["spec"] else "-"
        lines.append(
            f"| {item_counter} "
            f"| {item['name']} "
            f"| {spec_text} "
            f"| {item['qty']:g} "
            f"| {item['unit']} "
            f"| {client_up:,.0f} "
            f"| {client_total:,.0f} |"
        )
        item_counter += 1

    # 最後一個類別收尾
    if current_category is not None:
        close_table(current_category, cat_total)

    # 總計
    if project_category in ["室內設計", "預售客變"]:
        management_fee = 0
    else:
        management_fee = grand_total * 0.1
        
    final_total = grand_total + management_fee
    
    lines.append(f"---")
    lines.append(f"")
    lines.append(f"## 合計總表")
    lines.append(f"")
    lines.append(f"| 項目 | 金額 |")
    lines.append(f"| :--- | ---: |")
    lines.append(f"| 合計金額 | NT$ {grand_total:,.0f} |")
    
    if management_fee > 0:
        lines.append(f"| 工程管理費 (10%) | NT$ {management_fee:,.0f} |")
        
    lines.append(f"| **總計金額** | **NT$ {final_total:,.0f}** |")
    lines.append(f"")
    
    if management_fee > 0:
        lines.append(f"> ※ 以上報價均為新台幣，內含10%工程管理費用。實際施工以現場丈量確認為主。")
    else:
        lines.append(f"> ※ 以上報價均為新台幣。實際施工以現場丈量確認為主。")


    md_content = "\n".join(lines)
    with open(output_path, "w", encoding="utf-8") as f:
        f.write(md_content)

    print(f"  [SAVED] Markdown：{output_path.name}")
    return md_content


# ─── 主流程 ───────────────────────────────────────────────
def main():
    print("\n" + "="*60)
    print("  JYY DESIGN 報價精算系統")
    print("="*60)

    # 讀取設定檔
    config_path = BASE_DIR / "config.json"
    if config_path.exists():
        with open(config_path, encoding="utf-8") as f:
            config = json.load(f)
        markup_rate  = config.get("markup_rate", 1.3)
        project_name = config.get("project_name", "工程案件")
        client_name  = config.get("client_name",  "業主")
    else:
        # 建立預設設定
        config = {
            "markup_rate":  1.3,
            "project_name": "工程案件名稱",
            "client_name":  "業主姓名"
        }
        with open(config_path, "w", encoding="utf-8") as f:
            json.dump(config, f, ensure_ascii=False, indent=2)
        print(f"\n  ⚠️  已建立預設 config.json，請修改後重新執行！")
        print(f"     路徑：{config_path}")
        markup_rate  = 1.3
        project_name = "工程案件"
        client_name  = "業主"

    print(f"  案件：{project_name} | 業主：{client_name}")
    print(f"  利潤率：{(markup_rate-1)*100:.0f}% (x{markup_rate})")
    print(f"  掃描：{INPUT_DIR}")

    vendors = scan_input_folder()

    if not vendors:
        print(f"\n  ❗ 輸入資料夾沒有可讀取的報價單！")
        print(f"     請將廠商報價 Excel 或 JSON 放入：{INPUT_DIR}")
        return

    total_items = sum(len(v["items"]) for v in vendors)
    print(f"  [OK] 共讀取 {len(vendors)} 家廠商，{total_items} 個項目")

    # 產出審核表
    timestamp = datetime.now().strftime("%Y%m%d_%H%M")
    review_path = REVIEW_DIR / f"審核_彙整表_{project_name}_{timestamp}.xlsx"
    cost, client = export_review_excel(vendors, markup_rate, project_name, review_path)

    # 產出對業主報價 Excel
    quote_path = OUTPUT_DIR / f"報價單_{project_name}_{client_name}_{timestamp}.xlsx"
    grand_total = export_client_quote_excel(
        vendors, markup_rate, project_name, client_name, quote_path
    )

    # 產出 Markdown 報價單（Quote_Engine Skill 標準輸出）
    output_formats = config.get("output_format", ["excel", "markdown"])
    md_content = ""
    if "markdown" in output_formats:
        md_path = OUTPUT_DIR / f"報價單_{project_name}_{client_name}_{timestamp}.md"
        md_content = export_markdown_quote(
            vendors, markup_rate, project_name, client_name, md_path
        )

    print(f"  成本：NT$ {cost:,.0f} | 售價：NT$ {client:,.0f} | 利潤：NT$ {client-cost:,.0f}")


if __name__ == "__main__":
    main()
